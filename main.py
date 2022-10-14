from config import geo_token
from dadata import Dadata
from bs4 import BeautifulSoup
from selenium import webdriver
import time
from openpyxl import Workbook
from loguru import logger

def geocoder(city, geo_token):
    dadata = Dadata(geo_token)
    r = dadata.suggest("address", f"{city}")
    r1 = r[0]
    r2 = r1['data']
    global lon
    global lat
    lon = r2['geo_lon']
    lat = r2['geo_lat']
    print(lat, lon)
    return (lon, lat)


def get_sourse(url):

    logger.debug('Успешный запуск!')

    driver = webdriver.Chrome(
        executable_path='/usr/local/bin/chromedriver'
    )
    driver.maximize_window()

    driver.get(url=url)
    time.sleep(5)

    with open('weather_project/result.html', 'w') as file:
        file.write(driver.page_source)


def get_days(file_path):
    with open(file_path) as file:
        src = file.read()
    soup = BeautifulSoup(src, 'lxml')
    items_span = soup.find_all('strong', class_='forecast-details__day-number')
    global days
    days = []
    for item in items_span:
        item_span = item.get_text()
        days.append(str(item_span))  
    return days


def get_magnetic_field(file_path):
    with open(file_path) as file:
        src = file.read()
    soup = BeautifulSoup(src, 'lxml')
    items_dl = soup.find_all('dl', class_='forecast-fields')
    status_magnetic_field = []
    for item in items_dl:
        item_dl = item.get_text('|', strip='True')
        status_magnetic_field.append(str(item_dl))
    global smf
    smf = []
    for item in status_magnetic_field:
        i = item.replace('УФ-индекс|1,|низкий|', '').replace('УФ-индекс|1,|низкий', 'None').replace('УФ-индекс|0,|низкий', 'None')
        smf.append(i)
    return smf


def get_weather_phenomenon(file_path):
    with open(file_path) as file:
        src = file.read()
    soup = BeautifulSoup(src, 'lxml')
    items_td = soup.find_all('td', class_='weather-table__body-cell weather-table__body-cell_type_condition')
    global weather_phenomenon
    weather_phenomenon = []
    for item in items_td:
        item_td = item.get_text()
        weather_phenomenon.append(item_td)
    return weather_phenomenon


def get_times_of_day(file_path):
    with open(file_path) as file:
        src = file.read()
    soup = BeautifulSoup(src, 'lxml')
    items_div = soup.find_all('div', class_='weather-table__daypart')
    global times_of_day
    times_of_day = []
    for item in items_div:
        item_div = item.get_text()
        times_of_day.append(item_div)
    return times_of_day


def get_temperature(file_path):
    with open(file_path) as file:
        src = file.read()
    soup = BeautifulSoup(src, 'lxml')
    items_div = soup.find_all('div', class_='weather-table__temp')
    temperatures = []
    for item in items_div:
        item_div = item.get_text('|', strip='True')
        temperatures.append(item_div)
    res = []
    for item in temperatures:
        i = item.replace('+', '')
        j = i.split('|…|')
        res.append(j)
    global temps
    temps = []
    for i in res:
        temps.append(int(i[0]))
    return temps


def get_pressure(file_path):
    with open(file_path) as file:
        src = file.read()
    soup = BeautifulSoup(src, 'lxml')
    items_td = soup.find_all('td', class_='weather-table__body-cell weather-table__body-cell_type_air-pressure')
    global pressures
    pressures = []
    for item in items_td:
        item_td = item.get_text()
        pressures.append(int(item_td))
    return pressures


def get_humidity(file_path):
    with open(file_path) as file:
        src = file.read()
    soup = BeautifulSoup(src, 'lxml')
    items_td = soup.find_all('td', class_='weather-table__body-cell weather-table__body-cell_type_humidity')
    humidities = []
    for item in items_td:
        item_td = item.get_text()
        humidities.append(item_td)
    global hum
    hum = []
    for item in humidities:
        i = item.replace('%', '')
        hum.append(int(i))
    return hum


def split_list(alist, wanted_parts):
    length = len(alist)
    global res
    res = [alist[i*length // wanted_parts: (i+1)*length // wanted_parts]
           for i in range(wanted_parts)]
    print(res)
    return res


def get_sheet_weather(days, times_of_day, temps, pressures, hum, weather_phenomenon, smf, res):
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:AN1')
    ws['A1'] = 'Время суток'
    ws.append(times_of_day)
    ws.merge_cells('A3:AN3')
    ws['A3'] = 'Температура'
    ws.append(temps)
    ws.merge_cells('A5:AN5')
    ws['A5'] = 'Давление'
    ws.append(pressures)
    ws.merge_cells('A7:AN7')
    ws['A7'] = 'Влажность'
    ws.append(hum)
    ws.merge_cells('A9:AN9')
    ws['A9'] = 'Погодное явление'
    ws.append(weather_phenomenon)
    ws['AO3'] = 'Средняя температура'
    ws['AO4'] = '=SUM(A4:AN4)/40'
    max_number = max([res[0][0], res[0][1], res[0][2], res[0][3]])
    min_number = min([res[0][0], res[0][1], res[0][2], res[0][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('A11:D11')
        ws['A11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('A11:D11')
        ws['A11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('A11:D11')
        ws['A11'] = 'Не ожидается резких скачков атмосферного давления'
    max_number = max([res[1][0], res[1][1], res[1][2], res[1][3]])
    min_number = min([res[1][0], res[1][1], res[1][2], res[1][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('E11:H11')
        ws['E11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('E11:H11')
        ws['E11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('E11:H11')
        ws['E11'] = 'Не ожидается резких скачков атмосферного давления'
    max_number = max([res[2][0], res[2][1], res[2][2], res[2][3]])
    min_number = min([res[2][0], res[2][1], res[2][2], res[2][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('I11:L11')
        ws['I11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('I11:L11')
        ws['I11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('I11:L11')
        ws['I11'] = 'Не ожидается резких скачков атмосферного давления'
    max_number = max([res[3][0], res[3][1], res[3][2], res[3][3]])
    min_number = min([res[3][0], res[3][1], res[3][2], res[3][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('M11:P11')
        ws['M11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('M11:P11')
        ws['M11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('M11:P11')
        ws['M11'] = 'Не ожидается резких скачков атмосферного давления'
    max_number = max([res[4][0], res[4][1], res[4][2], res[4][3]])
    min_number = min([res[4][0], res[4][1], res[4][2], res[4][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('Q11:T11')
        ws['Q11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('Q11:T11')
        ws['Q11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('Q11:T11')
        ws['Q11'] = 'Не ожидается резких скачков атмосферного давления'
    max_number = max([res[5][0], res[5][1], res[5][2], res[5][3]])
    min_number = min([res[5][0], res[5][1], res[5][2], res[5][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('U11:X11')
        ws['U11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('U11:X11')
        ws['U11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('U11:X11')
        ws['U11'] = 'Не ожидается резких скачков атмосферного давления'
    max_number = max([res[6][0], res[6][1], res[6][2], res[6][3]])
    min_number = min([res[6][0], res[6][1], res[6][2], res[6][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('Y11:AB11')
        ws['Y11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('Y11:AB11')
        ws['Y11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('Y11:AB11')
        ws['Y11'] = 'Не ожидается резких скачков атмосферного давления'
    max_number = max([res[7][0], res[7][1], res[7][2], res[7][3]])
    min_number = min([res[7][0], res[7][1], res[7][2], res[7][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('AC11:AF11')
        ws['AC11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('AC11:AF11')
        ws['AC11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('AC11:AF11')
        ws['AC11'] = 'Не ожидается резких скачков атмосферного давления'
    max_number = max([res[8][0], res[8][1], res[8][2], res[8][3]])
    min_number = min([res[8][0], res[8][1], res[8][2], res[8][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('AG11:AJ11')
        ws['AG11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('AG11:AJ11')
        ws['AG11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('AG11:AJ11')
        ws['AG11'] = 'Не ожидается резких скачков атмосферного давления'
    max_number = max([res[9][0], res[9][1], res[9][2], res[9][3]])
    min_number = min([res[9][0], res[9][1], res[9][2], res[9][3]])
    if (max_number - min_number) >= 5 and (pressures.index(max_number) > pressures.index(min_number)):
        ws.merge_cells('AK11:AN11')
        ws['AK11'] = 'Ожидается резкое увеличение атмосферного давления'
    elif (max_number - min_number) >= 5 and (pressures.index(max_number) < pressures.index(min_number)):
        ws.merge_cells('AK11:AN11')
        ws['AK11'] = 'Ожидается резкое падение атмосферного давления'
    elif (max_number - min_number) < 5:
        ws.merge_cells('AK11:AN11')
        ws['AK11'] = 'Не ожидается резких скачков атмосферного давления'

    ws.merge_cells('A12:J12')
    ws['A12'] = 'Дни'
    ws.append(days)
    ws.append(smf)
    wb.save('weather_project/weather1.xlsx')

    logger.debug('Парсер закончил работу!')


def main():
    city = input('Введите город: ')
    geocoder(city, geo_token)
    logger.add('weather_project/log_weather.json', format='{time}, {level}, {message}', level='DEBUG', rotation='10 KB', compression='zip', serialize=True)
    get_sourse(url=f'https://yandex.ru/pogoda/details?lat={lat}&lon={lon}')
    get_days(file_path='/Users/ludmilaromanova/Dev/weather_project/result.html')
    get_magnetic_field(file_path='/Users/ludmilaromanova/Dev/weather_project/result.html')
    get_weather_phenomenon(file_path='/Users/ludmilaromanova/Dev/weather_project/result.html')
    get_times_of_day(file_path='/Users/ludmilaromanova/Dev/weather_project/result.html')
    get_temperature(file_path='/Users/ludmilaromanova/Dev/weather_project/result.html')
    get_pressure(file_path='/Users/ludmilaromanova/Dev/weather_project/result.html')
    get_humidity(file_path='/Users/ludmilaromanova/Dev/weather_project/result.html')
    split_list(alist=pressures, wanted_parts=10)
    get_sheet_weather(days=days, temps=temps, times_of_day=times_of_day, pressures=pressures, hum=hum, weather_phenomenon=weather_phenomenon, smf=smf, res=res)


if __name__=='__main__':
    main()
